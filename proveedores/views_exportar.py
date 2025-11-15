"""
Vistas para exportar órdenes de compra a PDF y Excel
"""
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .models import OrdenCompraProveedor


@login_required
def exportar_orden_excel(request, pk):
    """Exportar orden de compra a Excel"""
    orden = get_object_or_404(OrdenCompraProveedor, pk=pk)
    detalles = orden.detalles.select_related('producto').all()
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orden {orden.numero_orden}"
    
    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezado
    ws.merge_cells('A1:F1')
    ws['A1'] = 'ORDEN DE COMPRA A PROVEEDOR'
    ws['A1'].font = titulo_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Información de la orden
    row = 3
    ws[f'A{row}'] = 'Número de Orden:'
    ws[f'B{row}'] = orden.numero_orden
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'Fecha:'
    ws[f'B{row}'] = timezone.localtime(orden.fecha_orden).strftime('%d/%m/%Y %H:%M')
    
    row += 1
    ws[f'A{row}'] = 'Proveedor:'
    ws[f'B{row}'] = orden.proveedor.razon_social
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = 'RUC:'
    ws[f'B{row}'] = orden.proveedor.ruc
    
    row += 1
    ws[f'A{row}'] = 'Teléfono:'
    ws[f'B{row}'] = orden.proveedor.telefono or ''
    
    row += 1
    ws[f'A{row}'] = 'Email:'
    ws[f'B{row}'] = orden.proveedor.email or ''
    
    if orden.fecha_entrega_estimada:
        row += 1
        ws[f'A{row}'] = 'Fecha Entrega Estimada:'
        ws[f'B{row}'] = orden.fecha_entrega_estimada.strftime('%d/%m/%Y')
    
    # Tabla de productos
    row += 2
    headers = ['#', 'Código', 'Producto', 'Cantidad', 'Precio Unit.', 'Subtotal']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Detalles de productos
    row += 1
    for idx, detalle in enumerate(detalles, 1):
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=detalle.producto.codigo_principal).border = border
        ws.cell(row=row, column=3, value=detalle.producto.nombre).border = border
        ws.cell(row=row, column=4, value=float(detalle.cantidad)).border = border
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(detalle.precio_unitario)).border = border
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=6, value=float(detalle.subtotal)).border = border
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        row += 1
    
    # Totales
    row += 1
    ws.cell(row=row, column=5, value='Subtotal:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(orden.subtotal)).number_format = '$#,##0.00'
    
    row += 1
    ws.cell(row=row, column=5, value='IVA 15%:').font = Font(bold=True)
    ws.cell(row=row, column=6, value=float(orden.iva)).number_format = '$#,##0.00'
    
    row += 1
    ws.cell(row=row, column=5, value='TOTAL:').font = Font(bold=True, size=12)
    ws.cell(row=row, column=6, value=float(orden.total)).font = Font(bold=True, size=12)
    ws.cell(row=row, column=6).number_format = '$#,##0.00'
    
    # Observaciones
    if orden.observaciones:
        row += 2
        ws.cell(row=row, column=1, value='Observaciones:').font = Font(bold=True)
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=orden.observaciones)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Crear respuesta
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Orden_{orden.numero_orden}.xlsx'
    
    return response


@login_required
def exportar_orden_pdf(request, pk):
    """Exportar orden de compra a PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    
    orden = get_object_or_404(OrdenCompraProveedor, pk=pk)
    detalles = orden.detalles.select_related('producto').all()
    
    # Crear PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Título
    elements.append(Paragraph('ORDEN DE COMPRA A PROVEEDOR', title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Información de la orden
    info_data = [
        ['Número de Orden:', orden.numero_orden],
        ['Fecha:', timezone.localtime(orden.fecha_orden).strftime('%d/%m/%Y %H:%M')],
        ['Estado:', orden.get_estado_display()],
        ['', ''],
        ['Proveedor:', orden.proveedor.razon_social],
        ['RUC:', orden.proveedor.ruc],
        ['Teléfono:', orden.proveedor.telefono or ''],
        ['Email:', orden.proveedor.email or ''],
    ]
    
    if orden.fecha_entrega_estimada:
        info_data.insert(3, ['Fecha Entrega:', orden.fecha_entrega_estimada.strftime('%d/%m/%Y')])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Tabla de productos
    table_data = [['#', 'Código', 'Producto', 'Cant.', 'P. Unit.', 'Subtotal']]
    
    for idx, detalle in enumerate(detalles, 1):
        # Usar Paragraph para permitir wrap de texto en el nombre del producto
        nombre_producto = Paragraph(detalle.producto.nombre, styles['Normal'])
        
        table_data.append([
            str(idx),
            detalle.producto.codigo_principal or '',
            nombre_producto,  # Ahora permite múltiples líneas
            f"{detalle.cantidad:.2f}",
            f"${detalle.precio_unitario:.2f}",
            f"${detalle.subtotal:.2f}"
        ])
    
    # Totales
    table_data.extend([
        ['', '', '', '', 'Subtotal:', f"${orden.subtotal:.2f}"],
        ['', '', '', '', 'IVA 15%:', f"${orden.iva:.2f}"],
        ['', '', '', '', 'TOTAL:', f"${orden.total:.2f}"],
    ])
    
    productos_table = Table(table_data, colWidths=[0.4*inch, 1*inch, 3.5*inch, 0.8*inch, 1*inch, 1.2*inch])
    productos_table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        
        # Contenido
        ('FONTNAME', (0, 1), (-1, -4), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -4), 9),
        ('ALIGN', (0, 1), (0, -4), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -4), 'RIGHT'),
        ('VALIGN', (0, 1), (-1, -4), 'TOP'),  # Alineación superior para texto multilínea
        
        # Totales
        ('FONTNAME', (4, -3), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (4, -1), (-1, -1), 11),
        ('ALIGN', (4, -3), (-1, -1), 'RIGHT'),
        
        # Bordes
        ('BOX', (0, 0), (-1, -4), 1, colors.black),
        ('INNERGRID', (0, 0), (-1, -4), 0.5, colors.grey),
        ('LINEABOVE', (4, -3), (-1, -3), 1, colors.black),
        ('LINEABOVE', (4, -1), (-1, -1), 2, colors.black),
    ]))
    
    elements.append(productos_table)
    
    # Observaciones
    if orden.observaciones:
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph('<b>Observaciones:</b>', styles['Normal']))
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(orden.observaciones, styles['Normal']))
    
    # Generar PDF
    doc.build(elements)
    
    # Respuesta
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Orden_{orden.numero_orden}.pdf'
    response.write(pdf)
    
    return response
